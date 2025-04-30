# -*- coding: utf-8 -*-
"""
Created on Wed Apr 30 16:41:31 2025

@author: bansala4846
"""

# Part 1: Imports, Setup, Helper Functions, Data Handling

import streamlit as st
import pandas as pd
import numpy as np
import re # For column name normalization
import io
import base64
import json
import zipfile # For exporting multiple files
from pathlib import Path

# Plotting & Visualization
import plotly.express as px
import plotly.graph_objects as go
import matplotlib.pyplot as plt # SHAP plots often use matplotlib

# Machine Learning Core
from sklearn.model_selection import train_test_split, GridSearchCV, KFold
from sklearn.preprocessing import StandardScaler, OneHotEncoder, LabelEncoder
from sklearn.compose import ColumnTransformer
from sklearn.pipeline import Pipeline
from sklearn.linear_model import LogisticRegression
from sklearn.ensemble import RandomForestClassifier
from sklearn.cluster import KMeans
import xgboost as xgb

# Metrics & Evaluation
from sklearn.metrics import (
    accuracy_score, classification_report, confusion_matrix, roc_auc_score,
    roc_curve, precision_recall_curve, auc, f1_score, precision_score, recall_score
)

# Explainability
import shap

# PDF Generation (Optional - requires installation: pip install fpdf2)
try:
    from fpdf import FPDF
    FPDF_AVAILABLE = True
except ImportError:
    FPDF_AVAILABLE = False

# Excel Handling (Requires installation: pip install openpyxl)
import openpyxl # Required by pandas for reading/writing Excel files

# --- Constants & Configuration ---

# Define required columns for validation
REQUIRED_COLUMNS = ['ProjectID', 'InitialCostEstimate', 'InitialScheduleDays', 'ScopeChanges']
TARGET_VARIABLE = 'DerailmentRisk_Actual' # Define the target variable name

# Features eligible for Simulation Sliders (adjust as needed)
SIMULATION_FEATURES_NUMERIC = ['InitialCostEstimate', 'InitialScheduleDays', 'ScopeChanges']
SIMULATION_FEATURES_CATEGORICAL = ['ResourceAvailability', 'TechnicalComplexity', 'VendorReliability'] # Example

# Features for Clustering (adjust as needed)
CLUSTERING_FEATURES = ['CostVariancePerc', 'ScheduleVariancePerc', 'ScopeChanges', 'InitialCostEstimate'] # Example

# AutoML/GridSearchCV Parameter Grids (Example)
PARAM_GRID = {
    'LogisticRegression': {
        'classifier__C': [0.1, 1.0, 10.0],
        'classifier__solver': ['liblinear'] # Good for smaller datasets
    },
    'RandomForestClassifier': {
        'classifier__n_estimators': [100, 200],
        'classifier__max_depth': [10, 20, None],
        'classifier__min_samples_leaf': [3, 5]
    },
    'XGBClassifier': {
        'classifier__n_estimators': [100, 200],
        'classifier__learning_rate': [0.05, 0.1],
        'classifier__max_depth': [5, 7]
    }
}

# --- Page Configuration & Styling ---
st.set_page_config(
    page_title="RiskLens Pro - Enhanced",
    page_icon="🚀",
    layout="wide",
    initial_sidebar_state="expanded",
)

# Basic Theme Toggle (Conceptual - Streamlit themes are preferred if available)
# This is a simplified placeholder. Real dark mode needs more CSS or Streamlit's theme feature.
def apply_theme(theme):
    # Base styles (can be expanded significantly)
    base_css = """
    <style>
        /* Common styles */
        .stButton>button { border-radius: 8px; padding: 0.5rem 1rem; font-weight: bold; margin: 5px; }
        .stMetric { border-radius: 8px; padding: 15px 20px; margin-bottom: 10px; }
        h1, h2, h3 { font-weight: bold; }
        .stTabs [data-baseweb="tab-list"] button { border-radius: 8px 8px 0 0; font-weight: bold; border-bottom: none; }
        .stTabs [data-baseweb="tab-panel"] { padding-top: 25px; border: none; }
        .block-container { padding: 2rem 3rem 3rem 3rem !important; } /* Adjust main padding */
        [data-testid="stSidebar"] { padding: 1rem; }
    </style>
    """
    if theme == "dark":
        dark_css = """
        <style>
            body { color: #FAFAFA; background-color: #1E1E1E; }
            .stApp { background-color: #1E1E1E; }
            .stButton>button { border: 1px solid #E67300; background-color: #E67300; color: #FFFFFF; }
            .stButton>button:hover { background-color: #D06300; border-color: #D06300; }
            .stDownloadButton>button { background-color: #555; border-color: #555; color: white; }
            .stDownloadButton>button:hover { background-color: #666; border-color: #666; }
            h1, h2, h3, h4, h5, h6 { color: #FAFAFA; }
            h3 { border-bottom: 2px solid #E67300; }
            .stMetric { background-color: #2F2F2F; border: 1px solid #444; border-left: 5px solid #E67300; }
            .stMetric > label { color: #AAAAAA; }
            .stMetric > div[data-testid="stMetricValue"] { color: #FAFAFA; }
            .stTabs [data-baseweb="tab-list"] button[aria-selected="true"] { background-color: #E67300; color: white; }
            .stTabs [data-baseweb="tab-list"] button { color: #AAAAAA; background-color: #444; }
            .stTabs [data-baseweb="tab-list"] { border-bottom: 2px solid #E67300; }
            .stTabs [data-baseweb="tab-panel"] { background-color: #1E1E1E; }
            div[data-testid="stVerticalBlock"]>div[style*="flex-direction: column;"]>div[data-testid="stVerticalBlock"],
            div[data-testid="stVerticalBlock"]>div[style*="flex-direction: column;"]>div[data-testid="stHorizontalBlock"] {
                 border-radius: 8px !important; border: 1px solid #444 !important; background-color: #2F2F2F !important;
            }
            .stDataFrame { border-radius: 8px; overflow: hidden; } /* Basic dark theme for dataframe */
             /* Add more specific dark theme styles */
        </style>
        """
        st.markdown(base_css + dark_css, unsafe_allow_html=True)
    else: # Light theme (default)
        light_css = """
        <style>
            body { color: #333333; background-color: #F5F5F5; }
            .stApp { background-color: #F5F5F5; }
            .stButton>button { border: 1px solid #E67300; background-color: #E67300; color: white; }
            .stButton>button:hover { background-color: #D06300; border-color: #D06300; }
            .stDownloadButton>button { background-color: #646469; border-color: #646469; color: white; }
            .stDownloadButton>button:hover { background-color: #505055; border-color: #505055; }
            h1, h2 { color: #000000; }
            h3 { color: #E67300; border-bottom: 2px solid #E67300; }
            h4, h5, h6 { color: #646469; }
            .stMetric { background-color: #FFFFFF; border: 1px solid #D6D6D8; border-left: 5px solid #E67300; }
            .stMetric > label { color: #646469; }
            .stMetric > div[data-testid="stMetricValue"] { color: #000000; }
            .stTabs [data-baseweb="tab-list"] button[aria-selected="true"] { background-color: #E67300; color: white; }
            .stTabs [data-baseweb="tab-list"] button { color: #646469; background-color: #E0E0E0; }
            .stTabs [data-baseweb="tab-list"] { border-bottom: 2px solid #E67300; }
            .stTabs [data-baseweb="tab-panel"] { background-color: #F5F5F5; }
             div[data-testid="stVerticalBlock"]>div[style*="flex-direction: column;"]>div[data-testid="stVerticalBlock"],
            div[data-testid="stVerticalBlock"]>div[style*="flex-direction: column;"]>div[data-testid="stHorizontalBlock"] {
                 border-radius: 8px !important; border: 1px solid #D6D6D8 !important; background-color: #FFFFFF !important;
                 box-shadow: 0 2px 4px rgba(0,0,0,0.05) !important;
            }
            .stDataFrame { border-radius: 8px; overflow: hidden; }
             /* Add more specific light theme styles */
        </style>
        """
        st.markdown(base_css + light_css, unsafe_allow_html=True)

# --- Helper Functions ---

def normalize_column_name(col_name):
    """
    Normalizes column names: removes special chars, converts to PascalCase.
    Example: 'proj id', 'Project_ID', 'project id #' -> 'ProjectID'
    """
    if not isinstance(col_name, str):
        col_name = str(col_name)
    # Remove special characters except underscore
    s = re.sub(r'[^\w\s_]+', '', col_name)
    # Replace spaces/underscores with a single space for splitting
    s = re.sub(r'[\s_]+', ' ', s)
    # Capitalize words and join
    s = ''.join(word.capitalize() for word in s.split())
    return s

def smart_load_data(uploaded_files):
    """
    Loads data from multiple uploaded Excel/CSV files, concatenates them,
    normalizes column names, validates required columns, and generates features.
    """
    if not uploaded_files:
        return None, "No files uploaded."

    all_dfs = []
    errors = []
    required_cols_normalized = {normalize_column_name(col) for col in REQUIRED_COLUMNS}

    for file in uploaded_files:
        try:
            if file.name.endswith('.csv'):
                df = pd.read_csv(file)
            elif file.name.endswith(('.xls', '.xlsx')):
                # Make sure 'openpyxl' is installed
                df = pd.read_excel(file, engine='openpyxl')
            else:
                errors.append(f"Unsupported file type: {file.name}")
                continue

            # --- 1. Normalize Column Names ---
            original_columns = df.columns.tolist()
            df.columns = [normalize_column_name(col) for col in original_columns]
            normalized_columns = df.columns.tolist()
            # Create a mapping for reference if needed later
            # col_mapping = dict(zip(original_columns, normalized_columns))

            # --- 2. Validate Required Columns ---
            missing_cols = required_cols_normalized - set(normalized_columns)
            if missing_cols:
                errors.append(f"File '{file.name}' is missing required columns (after normalization): {', '.join(missing_cols)}")
                continue # Skip this file if essential columns are missing

            all_dfs.append(df)

        except Exception as e:
            errors.append(f"Error processing file {file.name}: {e}")

    if not all_dfs:
        return None, "No valid data could be loaded. Errors: " + "; ".join(errors)

    # Concatenate all valid dataframes
    combined_df = pd.concat(all_dfs, ignore_index=True)

    # --- 3. Automatically Generate Features ---
    try:
        # Ensure required columns are numeric, coerce errors
        for col in ['InitialCostEstimate', 'InitialScheduleDays']:
            if col in combined_df.columns:
                combined_df[col] = pd.to_numeric(combined_df[col], errors='coerce')
            else: # Should have been caught by validation, but double-check
                 errors.append(f"Critical Error: Column '{col}' not found after concatenation.")
                 # Handle this error appropriately, maybe return None or raise exception

        # Generate Actuals (Using simple factors for demonstration - replace with real logic if available)
        # If ActualCost/ActualScheduleDays are already present, use them. Otherwise, simulate.
        if 'ActualCost' not in combined_df.columns:
             # Simulate based on ScopeChanges and InitialEstimate (Example logic)
             cost_factor = 1 + (pd.to_numeric(combined_df.get('ScopeChanges', 0), errors='coerce').fillna(0) * 0.05) + np.random.normal(0, 0.08, len(combined_df))
             combined_df['ActualCost'] = combined_df['InitialCostEstimate'] * cost_factor.clip(0.8, 2.5) # Add some bounds
             st.info("Generated 'ActualCost' based on ScopeChanges and random variation.")
        else:
            combined_df['ActualCost'] = pd.to_numeric(combined_df['ActualCost'], errors='coerce')


        if 'ActualScheduleDays' not in combined_df.columns:
            schedule_factor = 1 + (pd.to_numeric(combined_df.get('ScopeChanges', 0), errors='coerce').fillna(0) * 0.07) + np.random.normal(0, 0.1, len(combined_df))
            combined_df['ActualScheduleDays'] = combined_df['InitialScheduleDays'] * schedule_factor.clip(0.8, 2.8)
            st.info("Generated 'ActualScheduleDays' based on ScopeChanges and random variation.")
        else:
             combined_df['ActualScheduleDays'] = pd.to_numeric(combined_df['ActualScheduleDays'], errors='coerce')

        # Calculate Variances (handle potential division by zero or NaN)
        combined_df['CostVariancePerc'] = ((combined_df['ActualCost'] - combined_df['InitialCostEstimate']) / combined_df['InitialCostEstimate'].replace(0, np.nan)) * 100
        combined_df['ScheduleVariancePerc'] = ((combined_df['ActualScheduleDays'] - combined_df['InitialScheduleDays']) / combined_df['InitialScheduleDays'].replace(0, np.nan)) * 100

        # Fill NaNs resulting from division by zero or missing inputs
        combined_df['CostVariancePerc'].fillna(0, inplace=True)
        combined_df['ScheduleVariancePerc'].fillna(0, inplace=True)

        # Generate Target Variable: DerailmentRisk_Actual
        cost_over_threshold = combined_df['CostVariancePerc'] > 10 # Cost overrun > 10%
        schedule_over_threshold = combined_df['ScheduleVariancePerc'] > 15 # Schedule overrun > 15%
        combined_df[TARGET_VARIABLE] = (cost_over_threshold | schedule_over_threshold).astype(int)

        # Add placeholder columns for predictions/probabilities if they don't exist
        if 'DerailmentRisk_Predicted_Prob' not in combined_df.columns:
            combined_df['DerailmentRisk_Predicted_Prob'] = np.nan
        if 'DerailmentRisk_Predicted' not in combined_df.columns:
             combined_df['DerailmentRisk_Predicted'] = pd.NA # Use pandas NA for integer type

    except Exception as e:
        errors.append(f"Error during feature generation: {e}")
        # Decide if this is fatal - maybe return partial data with warning?
        # For now, return None if feature generation fails critically
        return None, "Error during feature generation: " + "; ".join(errors)

    # --- Final Data Cleaning ---
    # Optional: Drop duplicates based on ProjectID, keeping the last entry
    if 'ProjectID' in combined_df.columns:
        combined_df = combined_df.drop_duplicates(subset=['ProjectID'], keep='last')

    # Convert categorical columns explicitly to 'category' dtype for efficiency
    for col in combined_df.select_dtypes(include=['object']).columns:
        if col != 'ProjectID' and col != 'ProjectName': # Keep IDs/Names as objects
             # Limit categories if cardinality is very high (optional)
             if combined_df[col].nunique() / len(combined_df) < 0.5: # Example threshold
                  combined_df[col] = combined_df[col].astype('category')

    status_message = f"Successfully loaded and processed data from {len(all_dfs)} file(s)."
    if errors:
        status_message += " Warnings/Errors: " + "; ".join(errors)

    return combined_df, status_message


def get_feature_names_and_types(df, target_col):
    """Identifies numerical and categorical features, excluding target and IDs."""
    cols_to_exclude = [target_col, 'ProjectID', 'ProjectName',
                       'ActualCost', 'ActualScheduleDays', 'CostVariancePerc', 'ScheduleVariancePerc', # Exclude generated actuals/variances
                       'DerailmentRisk_Predicted_Prob', 'DerailmentRisk_Predicted'] # Exclude prediction outputs
    features = [col for col in df.columns if col not in cols_to_exclude and not col.endswith('_Actual')]

    numerical_features = df[features].select_dtypes(include=np.number).columns.tolist()
    categorical_features = df[features].select_dtypes(include=['object', 'category']).columns.tolist()

    # Ensure no overlap and features exist
    numerical_features = [f for f in numerical_features if f in df.columns]
    categorical_features = [f for f in categorical_features if f in df.columns]

    return numerical_features, categorical_features

def df_to_csv_bytes(df):
    """Converts DataFrame to CSV bytes for download."""
    output = io.StringIO()
    df.to_csv(output, index=False)
    return output.getvalue().encode('utf-8')

def df_to_excel_bytes(df_dict):
    """Converts a dictionary of DataFrames to Excel bytes for download."""
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        for sheet_name, df in df_dict.items():
            df.to_excel(writer, sheet_name=sheet_name, index=False)
    return output.getvalue()

def create_download_zip(files_dict):
    """Creates a zip file containing multiple files provided as a dictionary {filename: bytes_content}."""
    zip_buffer = io.BytesIO()
    with zipfile.ZipFile(zip_buffer, 'a', zipfile.ZIP_DEFLATED, False) as zip_file:
        for file_name, data_bytes in files_dict.items():
            zip_file.writestr(file_name, data_bytes)
    return zip_buffer.getvalue()

# --- Placeholder Functions for Later Implementation ---
# These will be filled in subsequent parts

def get_model_pipeline(model_name, numerical_features, categorical_features, use_grid_search=False):
    """Creates ML pipeline (Preprocessing + Model). Handles GridSearchCV."""
    # To be implemented in Part 2
    pass

def train_model_and_explainers(df, features, numerical_features, categorical_features, target, train_size, random_state, model_name, threshold, use_grid_search=False):
    """Handles data prep, splitting, training, prediction, and SHAP explainer init."""
    # To be implemented in Part 2 (will integrate SHAP)
    pass

def plot_shap_summary(shap_values, X_processed, feature_names):
    """Generates SHAP summary plot."""
    # To be implemented in Part 2
    pass

def plot_shap_force_individual(explainer, shap_values, instance_index, X_processed, feature_names):
    """Generates SHAP force plot for a single instance."""
    # To be implemented in Part 2
    pass

def run_simulation(project_data, model_pipeline, simulation_params):
    """Runs 'what-if' simulation based on slider inputs."""
    # To be implemented in Part 3
    pass

def perform_clustering(df, features, n_clusters):
    """Performs KMeans clustering and returns cluster labels."""
    # To be implemented in Part 3
    pass

def plot_cluster_results(df, features, cluster_labels):
    """Visualizes clustering results."""
    # To be implemented in Part 3
    pass

def create_kpi_summary(df, target_col, pred_prob_col, threshold):
    """Generates KPI summary dictionary."""
    # To be implemented in Part 3
    pass

def plot_probability_histogram(df, pred_prob_col, threshold):
    """Plots histogram of predicted probabilities."""
    # To be implemented in Part 3
    pass

def plot_risk_heatmap(df, group_col1, group_col2, risk_col):
    """Plots heatmap of risk by categories."""
    # To be implemented in Part 3
    pass

def plot_bubble_chart(df, x_col, y_col, size_col, color_col):
    """Plots bubble chart (e.g., cost vs schedule variance)."""
    # To be implemented in Part 3
    pass

def compare_with_benchmark(current_data, benchmark_data):
    """Compares current portfolio data against a benchmark."""
    # To be implemented in Part 4
    pass

def export_insights_pdf(insights_data):
    """Exports key insights to a PDF file."""
    # To be implemented in Part 4
    pass

# --- Initialize Session State ---
default_state = {
    'app_theme': 'light',
    'project_data': None, # Main dataframe holding project info
    'benchmark_data': None, # Dataframe for benchmark comparison
    'data_load_status': "No data loaded.",
    'model_pipeline': None, # Trained pipeline object
    'shap_explainer': None, # SHAP explainer object
    'shap_values': None, # Calculated SHAP values for test set
    'X_test_processed': None, # Processed test features (for SHAP)
    'X_test_original_index': None, # Index of original test data
    'y_test': None, # True labels for test set
    'y_pred_prob': None, # Predicted probabilities for test set
    'feature_importance': None, # Feature importance df
    'numerical_features': [], # List of numerical feature names used in model
    'categorical_features': [], # List of categorical feature names used in model
    'features_processed_names': [], # Feature names after preprocessing (for SHAP)
    'clustering_results': None, # DataFrame with cluster labels
    'simulation_results': None, # Results from 'what-if' simulation
    'model_trained': False,
    # --- User Selections ---
    'selected_model': 'RandomForestClassifier',
    'use_automl': False,
    'train_test_split_ratio': 0.8,
    'prediction_threshold': 0.5,
    'random_state_seed': 42,
    'selected_features_cluster': CLUSTERING_FEATURES[:2] if len(CLUSTERING_FEATURES) >= 2 else CLUSTERING_FEATURES, # Default for clustering
    'num_clusters': 4,
    # --- Sidebar Navigation ---
    'current_step': 'Upload', # Tracks progress: Upload -> Transform -> Train -> Explore
}

# Initialize state variables if they don't exist
for key, value in default_state.items():
    if key not in st.session_state:
        st.session_state[key] = value

# Apply theme at the start
apply_theme(st.session_state.app_theme)

# --- Sidebar ---
st.sidebar.title("RiskLens Pro Navigator")
st.sidebar.markdown("---")

# Step-by-step Guide
st.sidebar.subheader("Workflow Steps")
steps = ["Upload", "Transform", "Train", "Explore"]
# Highlight current step (simple text indication for now)
for step in steps:
    if step == st.session_state.current_step:
        st.sidebar.markdown(f"**➡️ {step}**")
    else:
        st.sidebar.markdown(f"&nbsp;&nbsp;&nbsp;{step}")
st.sidebar.markdown("---")

# Theme Toggle
st.sidebar.subheader("⚙️ Settings")
current_theme = st.session_state.app_theme
new_theme = st.sidebar.radio("Select Theme", ('light', 'dark'),
                             index=('light', 'dark').index(current_theme),
                             key='theme_selector')
if new_theme != current_theme:
    st.session_state.app_theme = new_theme
    st.rerun() # Re-run to apply the new theme CSS

st.sidebar.markdown("---")

# --- Data Upload Section (Sidebar) ---
st.sidebar.subheader("1. Upload Data")
uploaded_files = st.sidebar.file_uploader(
    "Upload Project Files (Excel/CSV)",
    type=['csv', 'xls', 'xlsx'],
    accept_multiple_files=True,
    key="data_uploader_sidebar"
)

if uploaded_files:
    # Attempt to load data immediately upon upload
    if st.session_state.project_data is None or st.sidebar.button("Reload Uploaded Data"): # Load if no data or user clicks reload
        with st.spinner("Processing uploaded files..."):
            df, status = smart_load_data(uploaded_files)
            st.session_state.project_data = df
            st.session_state.data_load_status = status
            st.session_state.model_trained = False # Reset model status
            st.session_state.model_pipeline = None
            st.session_state.shap_explainer = None
            st.session_state.shap_values = None
            # Update current step if successful
            if df is not None:
                st.session_state.current_step = "Transform"
                st.sidebar.success("Data loaded. Proceed to Transform.")
                st.rerun() # Rerun to reflect changes and update UI elements
            else:
                st.sidebar.error(f"Data loading failed: {status}")

# Display data status
st.sidebar.info(st.session_state.data_load_status)
if st.session_state.project_data is not None:
    st.sidebar.success(f"✅ Data loaded ({len(st.session_state.project_data)} projects)")
    st.session_state.current_step = max(st.session_state.current_step, "Transform") # Ensure step is at least Transform

# --- Main Application Area (Title) ---
st.title("🚀 RiskLens Pro - Enhanced Project Risk Prediction")
st.markdown("_Leveraging advanced analytics for proactive project management._")
st.markdown("---")

# --- Placeholder for Tabs (to be defined in later parts) ---
# Example: tab1, tab2, tab3, tab4 = st.tabs(["📊 Dashboard", "⚙️ Model Training", "💡 Explainability", "🔬 Explore & Simulate"])

# Display initial message or data preview if loaded
if st.session_state.project_data is None:
    st.info("Welcome to RiskLens Pro! Please upload your project data using the sidebar to begin.")
else:
    st.success(f"Project data loaded. Navigate using the sidebar or the tabs above (once created). Current Step: **{st.session_state.current_step}**")
    if st.session_state.current_step == "Transform":
         st.markdown("### Data Transformation & Preview")
         st.markdown("The uploaded data has been processed. Key transformations include:")
         st.markdown(f"""
            * **Column Normalization:** Standardized column names (e.g., 'proj id' -> 'ProjectID').
            * **Required Columns:** Verified presence of {', '.join(REQUIRED_COLUMNS)}.
            * **Feature Generation:** Calculated 'CostVariancePerc', 'ScheduleVariancePerc', and the target '{TARGET_VARIABLE}'.
            * _(Note: 'ActualCost' and 'ActualScheduleDays' were generated using example logic if not present in the upload)._
         """)
         st.dataframe(st.session_state.project_data.head())
         st.markdown("---")
         st.info("Proceed to **Train** in the sidebar to configure and train a model.")


# End of Part 1

# <<< Part 1: Imports, Setup, Helper Functions, Data Handling (Code from previous response) >>>
# ... (Keep all code from Part 1 here) ...

# --- Machine Learning Pipeline & Training ---

# (Update the placeholder function from Part 1)
def get_model_pipeline(model_name, numerical_features, categorical_features, use_grid_search=False, random_state=42):
    """Creates ML pipeline (Preprocessing + Model). Handles GridSearchCV."""

    # Define transformers
    numeric_transformer = Pipeline(steps=[('scaler', StandardScaler())])
    categorical_transformer = Pipeline(steps=[('onehot', OneHotEncoder(handle_unknown='ignore', sparse_output=False))])

    # Define preprocessor
    preprocessor = ColumnTransformer(
        transformers=[
            ('num', numeric_transformer, numerical_features),
            ('cat', categorical_transformer, categorical_features)
        ],
        remainder='passthrough', # Keep other columns if any (shouldn't happen if features defined correctly)
        verbose_feature_names_out=False # Keep feature names simpler
    )
    # Ensure preprocessor outputs pandas DataFrame to retain feature names, crucial for SHAP
    preprocessor.set_output(transform="pandas")

    # Define base models
    models = {
        'LogisticRegression': LogisticRegression(random_state=random_state, class_weight='balanced', max_iter=2000, solver='liblinear'),
        'RandomForestClassifier': RandomForestClassifier(random_state=random_state, class_weight='balanced', n_jobs=-1),
        'XGBClassifier': xgb.XGBClassifier(random_state=random_state, use_label_encoder=False, eval_metric='logloss', objective='binary:logistic')
    }

    if model_name not in models:
        raise ValueError(f"Unsupported model_name: {model_name}. Choose from {list(models.keys())}")

    base_model = models[model_name]

    # Create the initial pipeline
    pipeline = Pipeline(steps=[('preprocessor', preprocessor), ('classifier', base_model)])

    # Wrap in GridSearchCV if requested
    if use_grid_search and model_name in PARAM_GRID:
        param_grid = PARAM_GRID[model_name]
        # Ensure parameter names in grid match pipeline steps (e.g., 'classifier__C')
        if not any(key.startswith('classifier__') for key in param_grid):
             st.warning(f"Parameter grid for {model_name} might be missing 'classifier__' prefix. Adjusting.")
             param_grid = {f"classifier__{k}": v for k, v in param_grid.items()}

        # Use cross-validation within GridSearchCV
        cv = KFold(n_splits=3, shuffle=True, random_state=random_state) # Reduced folds for speed
        grid_search = GridSearchCV(pipeline, param_grid, cv=cv, scoring='roc_auc', n_jobs=-1, error_score='raise')
        # Return the GridSearchCV object itself, it acts like an estimator
        return grid_search
    else:
        # Return the standard pipeline if not using GridSearchCV or no grid defined
        return pipeline

# (Update the placeholder function from Part 1)
@st.cache_resource(show_spinner="Training model and initializing explainers...") # Cache pipeline/explainers
def train_model_and_explainers(_df, _numerical_features, _categorical_features, target, train_size, random_state, model_name, threshold, use_grid_search=False):
    """Handles data prep, splitting, training, prediction, and SHAP explainer init."""
    results = {'success': False, 'message': '', 'model_pipeline': None, 'shap_explainer': None, 'shap_values': None,
               'X_test_processed': None, 'X_test_original_index': None, 'y_test': None, 'y_pred_prob': None,
               'feature_importance': None, 'features_processed_names': None, 'best_params': None,
               'preprocessor': None} # Added preprocessor to results

    features = _numerical_features + _categorical_features
    if not features:
        results['message'] = "❌ Error: No features selected for training."
        return results

    try:
        # Prepare data
        X = _df[features].copy()
        y = _df[target].copy()

        # Drop rows where target is NaN
        valid_indices = y.dropna().index
        X = X.loc[valid_indices]
        y = y.loc[valid_indices]
        y = y.astype(int) # Ensure target is integer

        if X.empty or y.empty:
            results['message'] = "❌ Error: No valid data remaining after removing rows with missing target values."
            return results
        if len(X) < 10: # Need sufficient samples
             results['message'] = f"❌ Error: Insufficient data ({len(X)} samples) for training after cleaning."
             return results

        # Impute missing values in features (simple mean/mode imputation)
        # Consider more sophisticated imputation if needed
        for col in _numerical_features:
            if X[col].isnull().any():
                mean_val = X[col].mean()
                if pd.isna(mean_val): mean_val = 0
                X[col].fillna(mean_val, inplace=True)
        for col in _categorical_features:
            if X[col].isnull().any():
                mode_val = X[col].mode()
                fill_val = mode_val[0] if not mode_val.empty else 'Unknown'
                X[col].fillna(fill_val, inplace=True)
                X[col] = X[col].astype(str) # Ensure categorical are string type before OHE


        # Split data (handle potential stratification issues)
        try:
            X_train, X_test, y_train, y_test = train_test_split(
                X, y, train_size=train_size, random_state=random_state, stratify=y
            )
        except ValueError:
            st.warning("Stratification failed (likely due to small class size). Splitting without stratification.")
            if len(X) < 2: # Need at least 2 samples for train/test split
                 results['message'] = f"❌ Error: Dataset too small ({len(X)} samples) to create a train/test split."
                 return results
            # Adjust train_size if dataset is very small to ensure test set has at least 1 sample
            min_test_samples = 1
            max_train_size = 1.0 - (min_test_samples / len(X)) if len(X) > min_test_samples else 0.0
            adjusted_train_size = min(train_size, max_train_size) if len(X) > min_test_samples else train_size
            if adjusted_train_size <= 0 or (1.0-adjusted_train_size)*len(X) < min_test_samples:
                results['message'] = f"❌ Error: Dataset too small ({len(X)} samples) to create a valid train/test split."
                return results

            X_train, X_test, y_train, y_test = train_test_split(
                X, y, train_size=adjusted_train_size, random_state=random_state
            )


        # --- Get and Train Pipeline ---
        pipeline_object = get_model_pipeline(model_name, _numerical_features, _categorical_features, use_grid_search, random_state)
        pipeline_object.fit(X_train, y_train)

        # If GridSearchCV was used, get the best estimator
        if use_grid_search and isinstance(pipeline_object, GridSearchCV):
            final_pipeline = pipeline_object.best_estimator_
            results['best_params'] = pipeline_object.best_params_
            best_params_str = ", ".join([f"{k.split('__')[1]}={v}" for k, v in results['best_params'].items()])
            results['message'] = f"✅ AutoML ({model_name}) trained successfully! Best Params: {best_params_str}."
        else:
            final_pipeline = pipeline_object
            results['message'] = f"✅ {model_name} model trained successfully!"

        # Extract preprocessor and trained classifier
        preprocessor = final_pipeline.named_steps['preprocessor']
        classifier = final_pipeline.named_steps['classifier']
        results['preprocessor'] = preprocessor # Store the fitted preprocessor

        # Transform test data *using the fitted preprocessor*
        X_test_processed = preprocessor.transform(X_test)
        # Get feature names after processing (important for SHAP)
        processed_feature_names = preprocessor.get_feature_names_out().tolist()

        # --- Predictions ---
        y_pred_prob = final_pipeline.predict_proba(X_test)[:, 1]

        # --- Feature Importance ---
        importance_df = None
        try:
            if hasattr(classifier, 'feature_importances_'): # Tree-based models
                importances = classifier.feature_importances_
            elif hasattr(classifier, 'coef_'): # Linear models
                importances = np.abs(classifier.coef_[0])
            else:
                importances = None # Model type might not have standard importance

            if importances is not None and len(processed_feature_names) == len(importances):
                importance_df = pd.DataFrame({'Feature': processed_feature_names, 'Importance': importances})
                importance_df = importance_df.sort_values(by='Importance', ascending=False).reset_index(drop=True)
            elif importances is not None:
                st.warning(f"Feature importance length mismatch: {len(processed_feature_names)} names vs {len(importances)} importances.")
                # Fallback: Maybe just show coefficients/importances without names if mismatch occurs
                importance_df = pd.DataFrame({'Importance': importances})

        except Exception as e:
            st.warning(f"Could not calculate feature importance: {e}")

        # --- SHAP Explainer Initialization ---
        shap_explainer = None
        shap_values = None
        try:
            # Transform train data for background dataset (can be slow for KernelExplainer)
            X_train_processed = preprocessor.transform(X_train) # Must use fitted preprocessor

            # Select appropriate SHAP explainer based on the *final* classifier type
            if isinstance(classifier, (RandomForestClassifier, xgb.XGBClassifier)):
                st.write("Using TreeExplainer for SHAP...")
                explainer = shap.TreeExplainer(classifier, data=X_train_processed, feature_perturbation="interventional", model_output="probability") # Using probability output
            elif isinstance(classifier, LogisticRegression):
                 st.write("Using LinearExplainer for SHAP...")
                 # Linear explainer expects model and data (masker often used)
                 # It might need specific setup depending on data correlation assumptions
                 masker = shap.maskers.Independent(data=X_train_processed)
                 explainer = shap.LinearExplainer(classifier, masker=masker, model_output="probability") # Check if model_output works
            else:
                # Fallback to KernelExplainer (can be slow!)
                st.warning("Using KernelExplainer for SHAP (can be slow). Consider specific explainers if possible.")
                 # For KernelExplainer, need a function that takes processed numpy array and returns probabilities
                def predict_proba_func(X_proc_arr):
                    X_proc_df = pd.DataFrame(X_proc_arr, columns=processed_feature_names)
                    # Need to ensure the predict function handles numpy correctly if pipeline expects DataFrame
                    # May need reconstruction if pipeline internally converts types
                    try:
                        probs = final_pipeline.predict_proba(X_proc_df)[:, 1]
                        return probs
                    except Exception as e:
                        # Attempt prediction without DataFrame conversion if first fails
                        try:
                             return final_pipeline.predict_proba(X_proc_arr)[:, 1]
                        except:
                             st.error(f"KernelExplainer predict function error: {e}")
                             # Return dummy values to avoid crashing, though explanations will be wrong
                             return np.zeros(X_proc_arr.shape[0])


                # Sample background data for performance
                background_data_sample = shap.sample(X_train_processed, 100) if len(X_train_processed) > 100 else X_train_processed
                explainer = shap.KernelExplainer(predict_proba_func, background_data_sample)

            # Calculate SHAP values for the test set (can take time)
            # Ensure X_test_processed is passed correctly (DataFrame or Numpy depending on explainer)
            st.write(f"Calculating SHAP values on {X_test_processed.shape[0]} test instances...")
            shap_values = explainer.shap_values(X_test_processed, check_additivity=False) # check_additivity=False can prevent errors but check validity

            # For binary classification, shap_values might return a list [shap_values_class0, shap_values_class1]
            # Or just shap_values_class1 depending on explainer/model. We typically want class 1 (positive class).
            if isinstance(shap_values, list) and len(shap_values) == 2:
                shap_values = shap_values[1] # Use SHAP values for the positive class (Derailment=1)

            results['shap_explainer'] = explainer
            results['shap_values'] = shap_values
            st.write("SHAP calculation complete.")

        except Exception as e_shap:
            st.error(f"❌ SHAP Initialization/Calculation Error: {e_shap}")
            results['message'] += " (SHAP failed)"

        # Store results
        results['success'] = True
        results['model_pipeline'] = final_pipeline # Store the best pipeline (GridSearchCV or regular)
        results['X_test_processed'] = X_test_processed
        results['X_test_original_index'] = X_test.index # Store the original index
        results['y_test'] = y_test
        results['y_pred_prob'] = y_pred_prob
        results['feature_importance'] = importance_df
        results['features_processed_names'] = processed_feature_names


    except MemoryError:
        results['message'] = "❌ Memory Error during training/SHAP. Try reducing data size or complexity."
    except Exception as e:
        results['message'] = f"❌ Error during training: {e}"
        st.exception(e) # Log the full traceback for debugging

    return results


# --- SHAP Plotting Functions ---

# (Update the placeholder function from Part 1)
def plot_shap_summary(shap_values, X_processed, feature_names=None, plot_type="dot"):
    """Generates SHAP summary plot and displays in Streamlit."""
    try:
        fig, ax = plt.subplots()
        # Ensure X_processed is a DataFrame with correct column names if feature_names not provided
        if feature_names is None and isinstance(X_processed, pd.DataFrame):
            feature_names = X_processed.columns.tolist()

        if isinstance(X_processed, pd.DataFrame):
             shap.summary_plot(shap_values, X_processed, feature_names=feature_names, show=False, plot_type=plot_type)
        else: # If X_processed is numpy
             shap.summary_plot(shap_values, X_processed, feature_names=feature_names, show=False, plot_type=plot_type)

        plt.title("SHAP Feature Importance Summary")
        plt.tight_layout()
        st.pyplot(fig)
        plt.close(fig) # Close the plot to free memory
    except Exception as e:
        st.error(f"Error generating SHAP summary plot: {e}")

# (Update the placeholder function from Part 1)
def plot_shap_force_individual(explainer, shap_values, instance_index, X_processed, feature_names=None):
    """Generates SHAP force plot for a single instance."""
    try:
        # Check if the explainer has the expected_value attribute
        if not hasattr(explainer, 'expected_value'):
            st.error("SHAP explainer does not have 'expected_value'. Cannot generate force plot.")
            # Try calculating expected value if possible (might be specific to explainer type)
            # Example for TreeExplainer: explainer.expected_value = explainer.model.predict(shap.sample(X_background, 100)).mean()
            return # Exit if expected_value is missing

        # Handle shap_values potentially being list (for multi-class or certain explainers)
        sv_instance = shap_values
        if isinstance(shap_values, list): # Should have been handled in train function, but double check
             sv_instance = shap_values[1] if len(shap_values) > 1 else shap_values[0]

        # Get the expected value for the positive class if it's an array/list
        expected_val = explainer.expected_value
        if isinstance(expected_val, (np.ndarray, list)) and len(expected_val) > 1:
            expected_val = expected_val[1] # Assuming index 1 is the positive class

        # Select the instance data
        instance_data = X_processed.iloc[[instance_index]] if isinstance(X_processed, pd.DataFrame) else X_processed[instance_index:instance_index+1]

        # Generate the plot (use matplotlib version for Streamlit compatibility)
        st.markdown("##### SHAP Force Plot (Local Explanation)")
        st.info("This plot shows features pushing the prediction higher (red) or lower (blue) than the baseline (expected value).")
        shap.force_plot(expected_val,
                        sv_instance[instance_index,:],
                        instance_data, # Pass the features of the instance
                        feature_names=feature_names, # Pass feature names if available
                        matplotlib=True, show=False)
        st.pyplot(plt.gcf(), bbox_inches='tight', dpi=150) # Use plt.gcf() to get the current figure
        plt.clf() # Clear the current matplotlib figure
        plt.close() # Ensure plot is closed

    except IndexError:
         st.error(f"Error generating SHAP force plot: Instance index {instance_index} out of bounds for SHAP values.")
    except AttributeError as ae:
         st.error(f"SHAP Force Plot Error (AttributeError): {ae}. Check explainer type and expected_value.")
    except Exception as e:
        st.error(f"Error generating SHAP force plot: {e}")
        st.exception(e)


# --- Sidebar Updates (Model Config & Training) ---

st.sidebar.subheader("2. Model Configuration")

# Select Model
model_options = ['LogisticRegression', 'RandomForestClassifier', 'XGBClassifier']
st.session_state.selected_model = st.sidebar.selectbox(
    "Select Model Algorithm:", model_options,
    index=model_options.index(st.session_state.selected_model) if st.session_state.selected_model in model_options else 0,
    key="model_select_sidebar"
)

# AutoML Toggle
st.session_state.use_automl = st.sidebar.checkbox(
    "Tune Hyperparameters (AutoML/GridSearch)",
    value=st.session_state.use_automl,
    key="automl_toggle_sidebar",
    help=f"If checked, uses GridSearchCV with predefined parameters ({', '.join(PARAM_GRID.keys())}) to find the best model settings. Training will take longer."
)

# Prediction Threshold
st.session_state.current_prediction_threshold = st.sidebar.slider(
    "Prediction Threshold:", min_value=0.01, max_value=0.99,
    value=st.session_state.current_prediction_threshold, step=0.01,
    key="pred_threshold_sidebar", help="Probability threshold to classify as 'High Risk'. Affects metrics like Accuracy, Precision, Recall, F1."
)

# Feature Selection & Parameters Expander
with st.sidebar.expander("Select Features & Parameters"):
    if st.session_state.project_data is not None:
        df_sidebar_config = st.session_state.project_data
        potential_num_features, potential_cat_features = get_feature_names_and_types(df_sidebar_config, TARGET_VARIABLE)

        # Use existing state if available, otherwise default to all potential features
        default_num = st.session_state.numerical_features if st.session_state.numerical_features else potential_num_features
        default_cat = st.session_state.categorical_features if st.session_state.categorical_features else potential_cat_features

        # Filter defaults to only include available columns
        default_num = [f for f in default_num if f in potential_num_features]
        default_cat = [f for f in default_cat if f in potential_cat_features]

        st.session_state.numerical_features = st.multiselect(
            "Select Numerical Features:", potential_num_features,
            default=default_num,
            key="num_features_sidebar_select"
        )
        st.session_state.categorical_features = st.multiselect(
            "Select Categorical Features:", potential_cat_features,
            default=default_cat,
            key="cat_features_sidebar_select"
        )

        # Train/test split settings
        st.session_state.train_test_split_ratio = st.slider(
            "Training Data Proportion:", 0.5, 0.9,
            st.session_state.train_test_split_ratio, 0.05,
            key="train_split_sidebar_slider"
            )
        st.session_state.random_state_seed = st.number_input(
            "Random Seed:", 0, 1000,
            st.session_state.random_state_seed, 1,
            key="random_seed_sidebar_input"
            )
    else:
        st.info("Load project data first (Step 1).")

st.sidebar.markdown("---")

# --- Train Button (Sidebar) ---
st.sidebar.subheader("3. Train Model")
train_button_sidebar = st.sidebar.button("🚀 Train Selected Model", key="train_model_sidebar", use_container_width=True)

if train_button_sidebar:
    selected_features = st.session_state.numerical_features + st.session_state.categorical_features
    if st.session_state.project_data is None:
        st.sidebar.error("❌ Cannot train: Project data not loaded.")
    elif not selected_features:
         st.sidebar.error("❌ Cannot train: No features selected in the expander.")
    elif TARGET_VARIABLE not in st.session_state.project_data.columns:
         st.sidebar.error(f"❌ Cannot train: Target '{TARGET_VARIABLE}' not found in data.")
    else:
        # Proceed with training
        model_display_name = f"AutoML ({st.session_state.selected_model})" if st.session_state.use_automl else st.session_state.selected_model

        training_results = train_model_and_explainers(
            _df=st.session_state.project_data,
            _numerical_features=st.session_state.numerical_features,
            _categorical_features=st.session_state.categorical_features,
            target=TARGET_VARIABLE,
            train_size=st.session_state.train_test_split_ratio,
            random_state=st.session_state.random_state_seed,
            model_name=st.session_state.selected_model,
            threshold=st.session_state.current_prediction_threshold,
            use_grid_search=st.session_state.use_automl
        )

        st.sidebar.info(training_results['message']) # Display message in sidebar

        if training_results['success']:
            # Update session state with all relevant results
            st.session_state.model_pipeline = training_results['model_pipeline']
            st.session_state.shap_explainer = training_results['shap_explainer']
            st.session_state.shap_values = training_results['shap_values']
            st.session_state.X_test_processed = training_results['X_test_processed']
            st.session_state.X_test_original_index = training_results['X_test_original_index']
            st.session_state.y_test = training_results['y_test']
            st.session_state.y_pred_prob = training_results['y_pred_prob']
            st.session_state.feature_importance = training_results['feature_importance']
            st.session_state.features_processed_names = training_results['features_processed_names']
            st.session_state.preprocessor = training_results['preprocessor'] # Store preprocessor
            # Note: Numerical/Categorical features used are already in session state

            # Update main dataframe predictions for the test set
            if st.session_state.X_test_original_index is not None and st.session_state.y_pred_prob is not None:
                # Ensure index exists in the main dataframe
                valid_test_indices = st.session_state.X_test_original_index.intersection(st.session_state.project_data.index)
                if len(valid_test_indices) == len(st.session_state.y_pred_prob):
                     # Clear previous predictions first
                    st.session_state.project_data['DerailmentRisk_Predicted_Prob'] = np.nan
                    st.session_state.project_data['DerailmentRisk_Predicted'] = pd.NA
                     # Assign new predictions only to the test set indices that are still valid
                    st.session_state.project_data.loc[valid_test_indices, 'DerailmentRisk_Predicted_Prob'] = st.session_state.y_pred_prob
                    st.session_state.project_data.loc[valid_test_indices, 'DerailmentRisk_Predicted'] = (st.session_state.y_pred_prob >= st.session_state.current_prediction_threshold).astype('Int64')
                    st.session_state.model_trained = True
                    st.session_state.current_step = "Explore" # Move to explore step
                    st.sidebar.success("Training complete. View results in tabs.")
                    # st.rerun() # Rerun might be needed if tabs depend heavily on immediate state update
                else:
                     st.sidebar.error(f"Prediction Error: Length mismatch or index issue.")
                     st.session_state.model_trained = False
            else:
                 st.sidebar.error("Prediction Error: Missing test index or probabilities.")
                 st.session_state.model_trained = False
        else:
            # Clear potentially inconsistent state if training failed
            st.session_state.model_pipeline = None
            st.session_state.shap_explainer = None
            st.session_state.shap_values = None
            st.session_state.model_trained = False

# --- Define Main Application Tabs ---
tab_titles = [
    "📊 Dashboard",
    "⚙️ Model Training & Explainability",
    "💡 Simulation & Clustering",
    "⚖️ Benchmarking",
    "📤 Export & Data",
]
# Check if model is trained to potentially enable more tabs or change titles
if st.session_state.model_trained:
    st.success(f"Model Trained: **{st.session_state.selected_model}{' (AutoML)' if st.session_state.use_automl else ''}**. Threshold: {st.session_state.current_prediction_threshold:.2f}. Proceed to **Explore**.")
else:
    st.info("Configure model parameters and click **'Train Selected Model'** in the sidebar.")

tabs = st.tabs(tab_titles)

# --- Tab 1: Dashboard (Placeholder) ---
with tabs[0]:
    st.header("📊 Portfolio Dashboard")
    st.markdown("_High-level overview of risk predictions and portfolio health._")
    st.info("Dashboard visuals (KPIs, Heatmap, Bubble Chart) will be implemented in Part 3.")
    if not st.session_state.model_trained:
         st.warning("Train a model first to populate the dashboard.")

# --- Tab 2: Model Training & Explainability ---
with tabs[1]:
    st.header("⚙️ Model Training & Explainability")
    st.markdown("_Evaluate model performance and understand prediction drivers._")

    if not st.session_state.model_trained:
        st.warning("Train a model using the sidebar (Step 3) to view results here.")
    elif st.session_state.y_test is None or st.session_state.y_pred_prob is None:
        st.error("Training finished, but test set results are missing. Please retrain.")
    else:
        # --- Model Performance Evaluation ---
        st.subheader("📊 Model Performance Evaluation (on Test Set)")
        try:
            y_true = st.session_state.y_test
            y_pred_prob = st.session_state.y_pred_prob
            y_pred_class = (y_pred_prob >= st.session_state.current_prediction_threshold).astype(int)

            accuracy = accuracy_score(y_true, y_pred_class)
            report_dict = classification_report(y_true, y_pred_class, output_dict=True, zero_division=0)
            report_df = pd.DataFrame(report_dict).transpose().round(3)
            cm = confusion_matrix(y_true, y_pred_class)

            roc_auc = None
            pr_auc = None
            if len(np.unique(y_true)) > 1: # AUC requires at least two classes
                 roc_auc = roc_auc_score(y_true, y_pred_prob)
                 precision, recall, _ = precision_recall_curve(y_true, y_pred_prob)
                 pr_auc = auc(recall, precision)

            f1 = f1_score(y_true, y_pred_class, zero_division=0)
            precision_s = precision_score(y_true, y_pred_class, zero_division=0)
            recall_s = recall_score(y_true, y_pred_class, zero_division=0)

            st.markdown(f"**Model:** {st.session_state.selected_model}{' (AutoML)' if st.session_state.use_automl else ''}")
            if st.session_state.use_automl and st.session_state.get('best_params'):
                 best_p_str = json.dumps(st.session_state['best_params'], indent=2)
                 st.code(f"Best Parameters (from GridSearchCV):\n{best_p_str}", language='json')

            # Display Metrics in columns
            col_m1, col_m2, col_m3, col_m4 = st.columns(4)
            col_m1.metric("Accuracy", f"{accuracy:.3f}")
            col_m2.metric("ROC AUC", f"{roc_auc:.3f}" if roc_auc is not None else "N/A")
            col_m3.metric("Precision", f"{precision_s:.3f}")
            col_m4.metric("Recall", f"{recall_s:.3f}")
            col_m1.metric("F1 Score", f"{f1:.3f}")
            col_m2.metric("PR AUC", f"{pr_auc:.3f}" if pr_auc is not None else "N/A")
            col_m3.metric("Threshold", f"{st.session_state.current_prediction_threshold:.2f}")
            col_m4.metric("Test Samples", f"{len(y_true)}")


            st.markdown("---")
            col_p1, col_p2 = st.columns(2)
            with col_p1:
                st.dataframe(report_df)
                st.caption("Classification Report (at chosen threshold)")

                # Confusion Matrix Plot (using Plotly helper from Part 1)
                fig_cm = plot_confusion_matrix_plotly(cm, labels=['Low Risk (0)', 'High Risk (1)'])
                st.plotly_chart(fig_cm, use_container_width=True)

            with col_p2:
                # ROC Curve Plot
                if roc_auc is not None:
                    fig_roc = plot_roc_curve_plotly(y_true, y_pred_prob, st.session_state.selected_model)
                    st.plotly_chart(fig_roc, use_container_width=True)
                else:
                    st.info("ROC curve cannot be calculated (requires at least two classes in test set).")

                 # Precision-Recall Curve Plot
                if pr_auc is not None:
                    fig_pr = plot_precision_recall_curve_plotly(y_true, y_pred_prob, st.session_state.selected_model)
                    st.plotly_chart(fig_pr, use_container_width=True)
                else:
                    st.info("PR curve cannot be calculated (requires at least two classes in test set).")

        except Exception as e:
            st.error(f"Error displaying model evaluation: {e}")
            st.exception(e)

        st.markdown("---")
        # --- SHAP Explainability ---
        st.subheader("💡 SHAP Explainability")

        if st.session_state.shap_explainer is None or st.session_state.shap_values is None or st.session_state.X_test_processed is None:
             st.warning("SHAP values were not calculated successfully during training. Explanations unavailable.")
        else:
            # Global Explanation (Summary Plot)
            st.markdown("##### Global Feature Importance (SHAP Summary)")
            st.info("Shows the impact of each feature on the model output. Higher SHAP values mean higher likelihood of 'High Risk'.")
            # Add toggle for plot type if desired
            shap_plot_type = st.selectbox("SHAP Summary Plot Type", ["dot", "bar"], index=0)
            plot_shap_summary(st.session_state.shap_values,
                              st.session_state.X_test_processed,
                              feature_names=st.session_state.features_processed_names,
                              plot_type=shap_plot_type)

            st.markdown("---")
            # Local Explanation (Force Plot for selected instance)
            st.markdown("##### Local Explanation (Individual Project)")
            st.markdown("Select a project **from the test set** to see what drove its specific prediction.")

            # Get Project IDs and Names for the test set
            test_indices = st.session_state.X_test_original_index
            if test_indices is not None and not test_indices.empty:
                test_projects_info = st.session_state.project_data.loc[test_indices, ['ProjectID', 'ProjectName']].reset_index()
                test_projects_info.rename(columns={'index': 'OriginalIndex'}, inplace=True) # Keep track of original index

                # Create mapping from original index to display string
                index_to_display = {row['OriginalIndex']: f"{row['ProjectID']} - {row['ProjectName']}" for _, row in test_projects_info.iterrows()}

                selected_original_index = st.selectbox(
                    "Select Test Project for Local Explanation:",
                    options=test_projects_info['OriginalIndex'].tolist(),
                    format_func=lambda x: index_to_display.get(x, f"Index {x}"),
                    key="project_explain_select"
                )

                if selected_original_index is not None:
                    # Find the corresponding index *within the SHAP values/X_test_processed array*
                    # This assumes the order is preserved from the original X_test split
                    try:
                        # Map original index back to its position in the test set array (0 to n_test_samples-1)
                        instance_array_index = test_indices.tolist().index(selected_original_index)

                        # Display Project Features
                        project_features = st.session_state.numerical_features + st.session_state.categorical_features
                        project_details = st.session_state.project_data.loc[selected_original_index, project_features]
                        st.dataframe(pd.DataFrame(project_details).T, use_container_width=True)

                        # Plot Force Plot
                        plot_shap_force_individual(st.session_state.shap_explainer,
                                                 st.session_state.shap_values,
                                                 instance_array_index,
                                                 st.session_state.X_test_processed,
                                                 feature_names=st.session_state.features_processed_names)
                    except ValueError:
                         st.error(f"Selected project index {selected_original_index} not found in the internal test set index mapping.")
                    except Exception as e_force:
                         st.error(f"Error displaying local explanation for project index {selected_original_index}: {e_force}")
            else:
                st.warning("Test set index information not available.")

# --- Tab 3: Simulation & Clustering (Placeholder) ---
with tabs[2]:
    st.header("💡 Simulation & Clustering")
    st.markdown("_Explore 'what-if' scenarios and discover risk driver patterns._")
    st.info("Simulation sliders and KMeans clustering visuals will be implemented in Part 3.")
    if not st.session_state.model_trained:
         st.warning("Train a model first to enable simulation and clustering.")

# --- Tab 4: Benchmarking (Placeholder) ---
with tabs[3]:
    st.header("⚖️ Benchmarking")
    st.markdown("_Compare your current portfolio against benchmark data._")
    st.info("Benchmark data upload and comparison logic will be implemented in Part 4.")

# --- Tab 5: Export & Data (Placeholder) ---
with tabs[4]:
    st.header("📤 Export & Data")
    st.markdown("_Download insights, data, or templates._")
    st.info("Export functionality (Excel/PDF/ZIP) and data download options will be implemented in Part 4.")
    # Add Download Templates section (can move from Data Management tab if preferred)
    st.subheader("Download Data Templates")
    with st.container():
        st.markdown("Download sample CSV templates for uploading data.")
        num_feat_template = st.session_state.numerical_features or SIMULATION_FEATURES_NUMERIC # Use simulation features as example
        cat_feat_template = st.session_state.categorical_features or SIMULATION_FEATURES_CATEGORICAL # Use simulation features as example
        required_template = list(REQUIRED_COLUMNS) + [TARGET_VARIABLE, 'ActualCost', 'ActualScheduleDays', 'CostVariancePerc', 'ScheduleVariancePerc']

        template_cols = {
            'Project Data': ['ProjectID', 'ProjectName'] + required_template + num_feat_template + cat_feat_template,
            # 'Risk Register': ['ProjectID', 'RiskID', ...], # Define if needed
            # 'Scenario Data': ['ProjectID', 'ScenarioName'] + SIMULATION_FEATURES_NUMERIC + SIMULATION_FEATURES_CATEGORICAL
            'Benchmark Data': ['MetricName', 'BenchmarkValue', 'PortfolioValue'] # Example benchmark structure
        }
        for data_type, cols in template_cols.items():
             # Ensure basic columns are present and remove duplicates
             unique_cols = pd.Index(cols).unique().tolist()
             template_df = pd.DataFrame(columns=unique_cols)
             csv_bytes = df_to_csv_bytes(template_df) # Use updated helper
             st.download_button(
                 label=f"Download {data_type} Template (.csv)", data=csv_bytes,
                 file_name=f"{data_type.lower().replace(' ', '_')}_template.csv", mime='text/csv',
                 key=f"download_template_{data_type.lower().replace(' ', '_')}_tab"
             )


# End of Part 2
# <<< Part 1: Imports, Setup, Helper Functions, Data Handling >>>
# ... (Keep all code from Part 1 here) ...

# <<< Part 2: Machine Learning Pipeline, Training, SHAP Integration, Sidebar Logic, Explainability Tab >>>
# ... (Keep all code from Part 2 here) ...

# --- Dashboard & Simulation/Clustering Functions ---

# (Update the placeholder function from Part 1)
def create_kpi_summary(df, pred_prob_col, pred_class_col):
    """Generates KPI summary dictionary."""
    kpis = {'Total Projects': 0, 'Projects with Prediction': 0, 'High-Risk Count': 0,
            'High-Risk Rate (%)': 0.0, 'Avg. Cost Overrun (%)': None, 'Avg. Schedule Overrun (%)': None,
            'Avg. Predicted Prob (%)': None}

    if df is None or df.empty:
        return kpis

    kpis['Total Projects'] = len(df)
    predicted_df = df.dropna(subset=[pred_prob_col, pred_class_col])
    kpis['Projects with Prediction'] = len(predicted_df)

    if kpis['Projects with Prediction'] > 0:
        kpis['High-Risk Count'] = int(predicted_df[pred_class_col].eq(1).sum())
        kpis['High-Risk Rate (%)'] = (kpis['High-Risk Count'] / kpis['Projects with Prediction']) * 100
        kpis['Avg. Predicted Prob (%)'] = predicted_df[pred_prob_col].mean() * 100

        # Add actual overrun KPIs if columns exist
        high_risk_actuals = predicted_df[predicted_df[pred_class_col] == 1]
        if 'CostVariancePerc' in df.columns:
             avg_cost_overrun = high_risk_actuals[high_risk_actuals['CostVariancePerc'] > 0]['CostVariancePerc'].mean()
             kpis['Avg. Cost Overrun (%)'] = avg_cost_overrun if pd.notna(avg_cost_overrun) else 0.0
        if 'ScheduleVariancePerc' in df.columns:
             avg_sched_overrun = high_risk_actuals[high_risk_actuals['ScheduleVariancePerc'] > 0]['ScheduleVariancePerc'].mean()
             kpis['Avg. Schedule Overrun (%)'] = avg_sched_overrun if pd.notna(avg_sched_overrun) else 0.0

    return kpis

# (Update the placeholder function from Part 1)
def plot_probability_histogram(df, pred_prob_col, threshold):
    """Plots histogram of predicted probabilities."""
    if df is None or df.empty or pred_prob_col not in df.columns or df[pred_prob_col].isnull().all():
        st.info("Prediction probability data not available for histogram.")
        return None

    fig = px.histogram(df.dropna(subset=[pred_prob_col]),
                       x=pred_prob_col, nbins=30,
                       title="Distribution of Predicted Risk Probabilities",
                       labels={pred_prob_col: 'Predicted Probability'},
                       opacity=0.75)
    fig.add_vline(x=threshold, line_dash="dash", line_color="red", annotation_text=f"Threshold ({threshold:.2f})")
    fig.update_layout(bargap=0.1, title_x=0.5, paper_bgcolor='rgba(0,0,0,0)', plot_bgcolor='rgba(0,0,0,0)')
    return fig

# (Update the placeholder function from Part 1)
def plot_risk_heatmap(df, group_col1, group_col2, risk_prob_col):
    """Plots heatmap of average risk probability by two categorical columns."""
    if df is None or df.empty or not all(c in df.columns for c in [group_col1, group_col2, risk_prob_col]) or df[risk_prob_col].isnull().all():
        st.info(f"Heatmap requires columns '{group_col1}', '{group_col2}', '{risk_prob_col}' with valid data.")
        return None

    try:
        heatmap_data = df.dropna(subset=[group_col1, group_col2, risk_prob_col])
        # Ensure groups are treated as strings/categories
        heatmap_data[group_col1] = heatmap_data[group_col1].astype(str)
        heatmap_data[group_col2] = heatmap_data[group_col2].astype(str)

        pivot = pd.pivot_table(heatmap_data, values=risk_prob_col, index=group_col1, columns=group_col2, aggfunc='mean')

        if pivot.empty:
             st.info(f"No data available to create heatmap for {group_col1} vs {group_col2}.")
             return None

        fig = go.Figure(data=go.Heatmap(
                       z=pivot.values,
                       x=pivot.columns.tolist(),
                       y=pivot.index.tolist(),
                       colorscale='OrRd', # Orange-Red scale for risk
                       zmin=0, zmax=1, # Probabilities range from 0 to 1
                       colorbar_title='Avg. Risk Prob.',
                       text=pivot.applymap(lambda x: f'{x:.1%}').values, # Format text as percentage
                       texttemplate="%{text}",
                       hoverongaps=False))
        fig.update_layout(
            title=f'Average Predicted Risk Probability by {group_col1} and {group_col2}',
            xaxis_title=group_col2, yaxis_title=group_col1,
            xaxis={'side': 'top'},
            yaxis_autorange='reversed', # Or keep default depending on preference
            paper_bgcolor='rgba(0,0,0,0)', plot_bgcolor='rgba(0,0,0,0)',
            margin=dict(l=20, r=20, t=60, b=20)
        )
        return fig
    except Exception as e:
        st.error(f"Error generating heatmap: {e}")
        return None

# (Update the placeholder function from Part 1)
def plot_bubble_chart(df, x_col, y_col, size_col, color_col, text_col='ProjectID'):
    """Plots bubble chart (e.g., cost vs schedule variance)."""
    required_cols = [x_col, y_col, size_col, color_col, text_col]
    if df is None or df.empty or not all(c in df.columns for c in required_cols) or df[color_col].isnull().all():
        st.info(f"Bubble chart requires columns: {', '.join(required_cols)} with valid data.")
        return None

    plot_df = df.dropna(subset=required_cols).copy()

    # Ensure size column is numeric and positive for bubble size
    plot_df[size_col] = pd.to_numeric(plot_df[size_col], errors='coerce')
    plot_df = plot_df[plot_df[size_col] > 0]

    if plot_df.empty:
        st.info(f"No valid data points for bubble chart after cleaning '{size_col}'.")
        return None

    # Normalize size for better visualization if needed (optional)
    # scaler = StandardScaler()
    # plot_df['size_normalized'] = scaler.fit_transform(plot_df[[size_col]]) + abs(scaler.fit_transform(plot_df[[size_col]]).min()) + 0.1 # Ensure positive

    fig = px.scatter(plot_df, x=x_col, y=y_col,
                     size=size_col, # Use original size, Plotly handles scaling
                     color=color_col,
                     hover_name=text_col,
                     hover_data={x_col:':.1f', y_col:':.1f', size_col:':,.0f', color_col:True}, # Customize hover info
                     color_continuous_scale=px.colors.sequential.OrRd, # Risk scale
                     title=f'{x_col} vs. {y_col} (Bubble Size: {size_col})')

    fig.update_layout(
        xaxis_title=x_col, yaxis_title=y_col,
        coloraxis_colorbar_title=color_col.replace('_',' '), # Nicer label
        title_x=0.5, paper_bgcolor='rgba(0,0,0,0)', plot_bgcolor='rgba(0,0,0,0)'
    )
    return fig

# (Update the placeholder function from Part 1)
@st.cache_data # Cache simulation result for the same inputs
def run_simulation(_base_project_data, _simulation_params, _model_pipeline):
    """Runs 'what-if' simulation based on slider inputs."""
    if _model_pipeline is None:
        return None, "Model not trained."
    if _base_project_data is None:
        return None, "Base project data not found."

    sim_data = _base_project_data.copy()
    features_used = list(_model_pipeline.named_steps['preprocessor'].feature_names_in_)

    for feature, value in _simulation_params.items():
        if feature in sim_data:
            sim_data[feature] = value
        elif feature in features_used: # Check if it's a feature the model expects
            st.warning(f"Simulation feature '{feature}' not in base data, but model expects it. Check data consistency.")
            # Potentially add the column with the value if necessary and possible
            # sim_data[feature] = value
        # else: ignore features not relevant to the model

    try:
        # Ensure sim_data is a DataFrame with the correct columns in the correct order
        sim_df = pd.DataFrame([sim_data], columns=features_used)

        # Need to handle potential NAs introduced or expected by model before prediction
        # Re-apply simple imputation consistent with training (or use fitted preprocessor if possible)
        # Note: Using the full pipeline's predict_proba automatically applies the fitted preprocessor
        pred_prob = _model_pipeline.predict_proba(sim_df)[0, 1] # Predict prob of class 1
        return pred_prob, "Simulation successful."

    except Exception as e:
        st.error(f"Error during simulation prediction: {e}")
        return None, f"Prediction Error: {e}"

# (Update the placeholder function from Part 1)
@st.cache_data # Cache clustering results
def perform_clustering(_df, _features, n_clusters, random_state):
    """Performs KMeans clustering and returns DataFrame with cluster labels."""
    if _df is None or _df.empty:
        return None, "Dataframe is empty."
    if not _features:
        return None, "No features selected for clustering."
    if n_clusters <= 1:
        return None, "Number of clusters must be greater than 1."

    cluster_data = _df[_features].copy()
    cluster_data.dropna(inplace=True) # Drop rows with NAs in selected features

    if len(cluster_data) < n_clusters:
        return None, f"Not enough data points ({len(cluster_data)}) for {n_clusters} clusters after dropping NAs."

    # Preprocessing: Scale numerical features
    num_cols_cluster = cluster_data.select_dtypes(include=np.number).columns
    if not num_cols_cluster.empty:
        scaler = StandardScaler()
        cluster_data[num_cols_cluster] = scaler.fit_transform(cluster_data[num_cols_cluster])

    # Perform KMeans
    try:
        kmeans = KMeans(n_clusters=n_clusters, init='k-means++', n_init=10, random_state=random_state)
        cluster_labels = kmeans.fit_predict(cluster_data)

        # Add labels back to the original index (handle dropped rows)
        _df_with_labels = _df.copy()
        # Create a temporary series with the original index of clustered data
        labels_series = pd.Series(cluster_labels, index=cluster_data.index, name='Cluster')
        # Merge/join back based on index
        _df_with_labels = _df_with_labels.join(labels_series)

        return _df_with_labels, f"KMeans clustering complete ({n_clusters} clusters)."
    except Exception as e:
        st.error(f"Error during KMeans clustering: {e}")
        return None, f"Clustering Error: {e}"

# (Update the placeholder function from Part 1)
def plot_cluster_results(df, features, cluster_col):
    """Visualizes clustering results using first two features."""
    if df is None or df.empty or cluster_col not in df.columns or not features or len(features) < 2:
        st.info("Cluster visualization requires dataframe with cluster labels and at least two features.")
        return None

    plot_df = df.dropna(subset=[features[0], features[1], cluster_col])
    plot_df[cluster_col] = plot_df[cluster_col].astype(str) # Treat cluster label as categorical for color

    if plot_df.empty:
        st.info("No data points available for cluster plot after removing NAs.")
        return None

    fig = px.scatter(plot_df, x=features[0], y=features[1], color=cluster_col,
                     title=f'Project Clusters based on {features[0]} and {features[1]}',
                     hover_name='ProjectID' if 'ProjectID' in plot_df.columns else None,
                     color_discrete_sequence=px.colors.qualitative.Plotly) # Use a qualitative color scale

    fig.update_layout(xaxis_title=features[0], yaxis_title=features[1], title_x=0.5,
                      paper_bgcolor='rgba(0,0,0,0)', plot_bgcolor='rgba(0,0,0,0)')
    fig.update_traces(marker=dict(size=8, opacity=0.7))
    return fig

# --- Populate Tab 1: Dashboard ---
with tabs[0]:
    st.header("📊 Portfolio Dashboard")
    st.markdown("_High-level overview of risk predictions and portfolio health._")

    if not st.session_state.model_trained:
         st.warning("Train a model first (Sidebar Step 3) to populate the dashboard.")
    elif st.session_state.project_data is None:
         st.warning("Project data not loaded.")
    else:
        df_dash = st.session_state.project_data

        # --- KPI Summary ---
        st.subheader("Key Performance Indicators")
        kpis = create_kpi_summary(df_dash, 'DerailmentRisk_Predicted_Prob', 'DerailmentRisk_Predicted')

        kpi_cols = st.columns(4)
        kpi_cols[0].metric("Total Projects", f"{kpis['Total Projects']}")
        kpi_cols[1].metric("Projects w/ Prediction", f"{kpis['Projects with Prediction']}")
        kpi_cols[2].metric("🚨 High-Risk Projects", f"{kpis['High-Risk Count']}")
        kpi_cols[3].metric("📈 High-Risk Rate", f"{kpis['High-Risk Rate (%)']:.1f}%")

        kpi_cols2 = st.columns(4)
        kpi_cols2[0].metric("Avg. Predicted Prob.", f"{kpis['Avg. Predicted Prob (%)']:.1f}%" if kpis['Avg. Predicted Prob (%)'] is not None else "N/A")
        # Display actual overruns only if available
        cost_overrun_display = f"{kpis['Avg. Cost Overrun (%)']:.1f}%" if kpis['Avg. Cost Overrun (%)'] is not None else "N/A"
        sched_overrun_display = f"{kpis['Avg. Schedule Overrun (%)']:.1f}%" if kpis['Avg. Schedule Overrun (%)'] is not None else "N/A"
        kpi_cols2[1].metric("Avg. Cost Overrun (HR)", cost_overrun_display, help="Average actual cost overrun for projects predicted as High Risk.")
        kpi_cols2[2].metric("Avg. Schedule Overrun (HR)", sched_overrun_display, help="Average actual schedule overrun for projects predicted as High Risk.")
        # Add more KPIs if needed

        st.markdown("---")
        # --- Visualizations ---
        st.subheader("Visual Overview")
        viz_cols = st.columns(2)

        with viz_cols[0]:
            # Probability Histogram
            fig_hist = plot_probability_histogram(df_dash, 'DerailmentRisk_Predicted_Prob', st.session_state.current_prediction_threshold)
            if fig_hist:
                st.plotly_chart(fig_hist, use_container_width=True)

            # Risk Heatmap
            # Choose default grouping columns if they exist, otherwise prompt user or skip
            group_col1_default = 'Region' if 'Region' in df_dash.columns else None
            group_col2_default = 'ProjectType' if 'ProjectType' in df_dash.columns else None
            if group_col1_default and group_col2_default:
                 fig_heatmap = plot_risk_heatmap(df_dash, group_col1_default, group_col2_default, 'DerailmentRisk_Predicted_Prob')
                 if fig_heatmap:
                     st.plotly_chart(fig_heatmap, use_container_width=True)
            else:
                 st.info("Cannot generate Risk Heatmap: Requires 'Region' and 'ProjectType' columns in the data.")

        with viz_cols[1]:
            # Bubble Chart (Cost vs Schedule Variance)
            # Define columns for bubble chart - ensure they exist
            x_col_bubble = 'CostVariancePerc'
            y_col_bubble = 'ScheduleVariancePerc'
            size_col_bubble = 'InitialCostEstimate'
            color_col_bubble = 'DerailmentRisk_Predicted_Prob'
            text_col_bubble = 'ProjectID'

            if all(c in df_dash.columns for c in [x_col_bubble, y_col_bubble, size_col_bubble, color_col_bubble]):
                fig_bubble = plot_bubble_chart(df_dash, x_col_bubble, y_col_bubble, size_col_bubble, color_col_bubble, text_col_bubble)
                if fig_bubble:
                    st.plotly_chart(fig_bubble, use_container_width=True)
            else:
                 missing_bubble_cols = [c for c in [x_col_bubble, y_col_bubble, size_col_bubble, color_col_bubble] if c not in df_dash.columns]
                 st.info(f"Cannot generate Bubble Chart: Missing required columns: {', '.join(missing_bubble_cols)}")


# --- Populate Tab 3: Simulation & Clustering ---
with tabs[2]:
    st.header("💡 Simulation & Clustering")
    st.markdown("_Explore 'what-if' scenarios and discover risk driver patterns._")

    if not st.session_state.model_trained:
         st.warning("Train a model first (Sidebar Step 3) to enable simulation and clustering.")
    elif st.session_state.project_data is None or st.session_state.model_pipeline is None:
         st.warning("Requires loaded data and a trained model.")
    else:
        df_sim = st.session_state.project_data
        pipeline = st.session_state.model_pipeline

        # --- What-If Simulation ---
        st.subheader("🎲 What-If Simulation")
        st.markdown("Select a project and adjust its features to see the impact on predicted risk.")

        project_list = df_sim['ProjectID'].tolist()
        project_name_map = df_sim.set_index('ProjectID')['ProjectName'].to_dict()
        selected_project_id_sim = st.selectbox(
            "Select Project for Simulation:", options=project_list,
            format_func=lambda x: f"{x} - {project_name_map.get(x, '')}",
            key="sim_project_select"
        )

        if selected_project_id_sim:
            base_project_data = df_sim[df_sim['ProjectID'] == selected_project_id_sim].iloc[0]
            original_prob = base_project_data['DerailmentRisk_Predicted_Prob']

            st.write(f"**Original Predicted Probability for {selected_project_id_sim}:** {original_prob:.1%}" if pd.notna(original_prob) else f"**Original Predicted Probability for {selected_project_id_sim}:** N/A (Not in test set or not predicted)")

            simulation_params = {}
            sim_cols = st.columns(2)

            # Sliders for Numeric Features
            with sim_cols[0]:
                st.markdown("**Adjust Numerical Features:**")
                for i, feat in enumerate(SIMULATION_FEATURES_NUMERIC):
                    if feat in base_project_data:
                         current_val = base_project_data[feat]
                         if pd.isna(current_val): # Handle missing base value
                             st.caption(f"{feat}: N/A in base data")
                             simulation_params[feat] = None # Or some default imputation?
                             continue
                         min_val = current_val * 0.5
                         max_val = current_val * 2.0
                         step = (max_val - min_val) / 20
                         # Ensure step is reasonable, e.g., integer for counts
                         if np.issubdtype(type(current_val), np.integer) or feat == 'ScopeChanges':
                              step = max(1, int(step))
                              min_val=int(min_val)
                              max_val=int(max_val) if max_val > min_val else int(min_val)+1
                              current_val = int(current_val)
                         else: # Float
                             step = float(step) if step > 0 else 1.0
                             current_val = float(current_val)

                         simulation_params[feat] = st.slider(f"{feat}", float(min_val), float(max_val), current_val, step=step, key=f"sim_slider_{feat}")
                    else:
                         st.caption(f"{feat}: Not in data")


            # Selectors for Categorical Features
            with sim_cols[1]:
                st.markdown("**Adjust Categorical Features:**")
                for feat in SIMULATION_FEATURES_CATEGORICAL:
                     if feat in base_project_data:
                         current_val = base_project_data[feat]
                         options = df_sim[feat].dropna().unique().tolist()
                         try:
                              # Ensure current_val is string for comparison if options are strings
                              current_val_str = str(current_val) if options and isinstance(options[0], str) else current_val
                              default_index = options.index(current_val_str) if current_val_str in options else 0
                         except ValueError:
                              default_index = 0
                         simulation_params[feat] = st.selectbox(f"{feat}", options, index=default_index, key=f"sim_select_{feat}")
                     else:
                          st.caption(f"{feat}: Not in data")

            # Run Simulation Prediction
            sim_prob, sim_status = run_simulation(base_project_data, pipeline, simulation_params)

            if sim_prob is not None:
                st.metric("Simulated Risk Probability", f"{sim_prob:.1%}", f"{sim_prob - (original_prob if pd.notna(original_prob) else 0):.1%}")
            else:
                st.error(f"Simulation failed: {sim_status}")
        else:
            st.info("Select a project to run a simulation.")


        st.markdown("---")
        # --- KMeans Clustering ---
        st.subheader("🧩 Risk Driver Clustering (KMeans)")
        st.markdown("Group projects based on selected risk drivers to identify common profiles.")

        # Allow user to select features for clustering
        potential_cluster_features = st.session_state.numerical_features + st.session_state.categorical_features # Or define specific list
        # Ensure features exist in the dataframe
        potential_cluster_features = [f for f in potential_cluster_features if f in df_sim.columns]

        # Use state for selected features
        st.session_state.selected_features_cluster = st.multiselect(
            "Select Features for Clustering:", potential_cluster_features,
            default=[f for f in st.session_state.selected_features_cluster if f in potential_cluster_features], # Filter default by available
            key="cluster_feature_select"
        )

        # Select number of clusters
        st.session_state.num_clusters = st.number_input(
            "Number of Clusters (K):", min_value=2, max_value=15,
            value=st.session_state.num_clusters, step=1,
            key="cluster_k_select"
        )

        if st.button("Run Clustering", key="run_clustering_button"):
            if not st.session_state.selected_features_cluster:
                 st.warning("Please select at least one feature for clustering.")
            else:
                with st.spinner(f"Running KMeans with K={st.session_state.num_clusters}..."):
                    df_clustered, cluster_status = perform_clustering(
                        _df=df_sim,
                        _features=st.session_state.selected_features_cluster,
                        n_clusters=st.session_state.num_clusters,
                        random_state=st.session_state.random_state_seed
                    )
                    st.info(cluster_status)
                    if df_clustered is not None:
                        st.session_state.clustering_results = df_clustered # Store clustered df

        # Display Clustering Results
        if st.session_state.clustering_results is not None:
            st.markdown("**Clustering Results**")
            df_display_clusters = st.session_state.clustering_results.copy()
            # Show dataframe with cluster labels
            display_cols_cluster = ['ProjectID', 'ProjectName'] + st.session_state.selected_features_cluster + ['Cluster']
            display_cols_cluster = [c for c in display_cols_cluster if c in df_display_clusters.columns] # Ensure columns exist
            st.dataframe(df_display_clusters[display_cols_cluster].dropna(subset=['Cluster']), use_container_width=True)

            # Visualize clusters (if 2+ features selected)
            if len(st.session_state.selected_features_cluster) >= 2:
                fig_clusters = plot_cluster_results(
                    st.session_state.clustering_results,
                    st.session_state.selected_features_cluster[:2], # Use first two selected features for plot axes
                    'Cluster'
                )
                if fig_clusters:
                    st.plotly_chart(fig_clusters, use_container_width=True)
                else:
                    st.info("Could not generate cluster plot (check data).")
            else:
                 st.info("Select at least two features to visualize clusters.")


# --- Tab 4: Benchmarking (Placeholder from Part 2) ---
# ... (Keep placeholder content) ...

# --- Tab 5: Export & Data (Placeholder & Templates from Part 2) ---
# ... (Keep placeholder content and download templates) ...

# End of Part 3

# <<< Part 1: Imports, Setup, Helper Functions, Data Handling >>>
# ... (Keep all code from Part 1 here) ...

# <<< Part 2: Machine Learning Pipeline, Training, SHAP Integration, Sidebar Logic, Explainability Tab >>>
# ... (Keep all code from Part 2 here) ...

# <<< Part 3: Dashboard Visuals, Simulation & Clustering Functions & Tabs >>>
# ... (Keep all code from Part 3 here) ...


# --- Benchmarking & Export Functions ---

# (Update the placeholder function from Part 1)
def compare_with_benchmark(current_kpis, benchmark_data):
    """Compares current portfolio KPIs against benchmark data."""
    if benchmark_data is None or benchmark_data.empty:
        return None, "Benchmark data not loaded or empty."
    if not all(c in benchmark_data.columns for c in ['MetricName', 'BenchmarkValue']):
         return None, "Benchmark data must contain 'MetricName' and 'BenchmarkValue' columns."

    comparison_list = []
    benchmark_dict = benchmark_data.set_index('MetricName')['BenchmarkValue'].to_dict()

    # Map KPI keys to potential MetricNames in benchmark file
    kpi_to_metric_name = {
        'High-Risk Rate (%)': 'High Risk Rate Pct',
        'Avg. Cost Overrun (%)': 'Avg Cost Overrun Pct (High Risk)',
        'Avg. Schedule Overrun (%)': 'Avg Schedule Overrun Pct (High Risk)',
        'Avg. Predicted Prob (%)': 'Avg Predicted Probability Pct',
        # Add mappings for other relevant KPIs if needed
    }

    for kpi_key, metric_name in kpi_to_metric_name.items():
        if kpi_key in current_kpis and current_kpis[kpi_key] is not None:
            portfolio_value = current_kpis[kpi_key]
            benchmark_value = benchmark_dict.get(metric_name)
            if benchmark_value is not None:
                try: # Ensure benchmark value is numeric for comparison
                     benchmark_value = float(benchmark_value)
                     diff = portfolio_value - benchmark_value
                     comparison_list.append({
                         'Metric': metric_name.replace(' Pct', ' (%)'), # Display friendly name
                         'Portfolio': portfolio_value,
                         'Benchmark': benchmark_value,
                         'Difference': diff
                     })
                except (ValueError, TypeError):
                     st.warning(f"Non-numeric benchmark value found for '{metric_name}'. Skipping comparison.")
            else:
                 st.warning(f"Benchmark metric '{metric_name}' not found in uploaded file.")


    if not comparison_list:
        return None, "No matching metrics found for comparison between portfolio KPIs and benchmark data."

    comparison_df = pd.DataFrame(comparison_list)
    return comparison_df, "Comparison complete."


# (Update the placeholder function from Part 1)
def export_insights_pdf(kpis, feature_importance_df, model_name, threshold, filename="RiskLensPro_Insights.pdf"):
    """Exports key insights (KPIs, top features) to a PDF file."""
    if not FPDF_AVAILABLE:
        st.error("PDF export requires the 'fpdf2' library. Please install it (`pip install fpdf2`).")
        return None

    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Helvetica", "B", 16)

    # Title
    pdf.cell(0, 10, "RiskLens Pro - Key Insights Report", 0, 1, "C")
    pdf.ln(10)

    # Summary Info
    pdf.set_font("Helvetica", "", 12)
    pdf.cell(0, 10, f"Model Used: {model_name}", 0, 1)
    pdf.cell(0, 10, f"Prediction Threshold: {threshold:.2f}", 0, 1)
    pdf.ln(5)

    # KPIs
    pdf.set_font("Helvetica", "B", 14)
    pdf.cell(0, 10, "Portfolio KPIs", 0, 1)
    pdf.set_font("Helvetica", "", 10)
    header_kpi = ['Metric', 'Value']
    col_width_kpi = pdf.w / 2.5 # Adjust width as needed
    # Print header
    for h in header_kpi:
         pdf.cell(col_width_kpi, 10, h, 1, 0, "C")
    pdf.ln()
    # Print data
    for key, value in kpis.items():
         display_value = f"{value:.1f}%" if '%' in key and isinstance(value, (int, float)) else f"{value}" if value is not None else "N/A"
         pdf.cell(col_width_kpi, 10, key, 1, 0)
         pdf.cell(col_width_kpi, 10, display_value, 1, 0)
         pdf.ln()
    pdf.ln(10)

    # Feature Importance
    if feature_importance_df is not None and not feature_importance_df.empty:
        pdf.set_font("Helvetica", "B", 14)
        pdf.cell(0, 10, "Top 10 Feature Importances", 0, 1)
        pdf.set_font("Helvetica", "", 10)
        header_feat = ['Feature', 'Importance']
        col_width_feat = pdf.w / 2.5
         # Print header
        for h in header_feat:
             pdf.cell(col_width_feat, 10, h, 1, 0, "C")
        pdf.ln()
         # Print data
        for _, row in feature_importance_df.head(10).iterrows():
             pdf.cell(col_width_feat, 10, str(row['Feature']), 1, 0)
             pdf.cell(col_width_feat, 10, f"{row['Importance']:.4f}", 1, 0)
             pdf.ln()
        pdf.ln(5)
        pdf.set_font("Helvetica", "I", 8)
        pdf.cell(0, 5, "Importance values depend on the model type (e.g., Gini impurity, coefficient magnitude, SHAP value).", 0, 1)

    # Save PDF to bytes buffer
    try:
        pdf_bytes = pdf.output(dest='S').encode('latin-1') # Output as bytes
        return pdf_bytes
    except Exception as e:
        st.error(f"Error generating PDF: {e}")
        return None


# --- Populate Tab 4: Benchmarking ---
with tabs[3]:
    st.header("⚖️ Benchmarking")
    st.markdown("_Compare your current portfolio against benchmark data._")

    if not st.session_state.model_trained:
        st.warning("Train a model first (Sidebar Step 3) to calculate current portfolio KPIs for benchmarking.")
    elif st.session_state.project_data is None:
        st.warning("Project data not loaded.")
    else:
        st.subheader("Upload Benchmark Data")
        st.info("Upload a CSV file with columns 'MetricName' and 'BenchmarkValue'. See template in 'Export & Data' tab.")
        benchmark_file = st.file_uploader("Upload Benchmark CSV", type=['csv'], key="benchmark_uploader")

        if benchmark_file:
            try:
                benchmark_df = pd.read_csv(benchmark_file)
                st.session_state.benchmark_data = benchmark_df # Store in session state
                st.success("Benchmark data uploaded successfully.")
                st.dataframe(benchmark_df.head(), use_container_width=True)
            except Exception as e:
                st.error(f"Error reading benchmark file: {e}")
                st.session_state.benchmark_data = None # Clear on error

        st.markdown("---")
        st.subheader("Performance vs. Benchmark")
        if st.session_state.benchmark_data is not None:
            # Recalculate current KPIs based on current data/predictions
            current_kpis_bench = create_kpi_summary(st.session_state.project_data, 'DerailmentRisk_Predicted_Prob', 'DerailmentRisk_Predicted')

            comparison_df, comparison_status = compare_with_benchmark(current_kpis_bench, st.session_state.benchmark_data)

            st.info(comparison_status)
            if comparison_df is not None:
                st.dataframe(comparison_df.style.format({
                    'Portfolio': '{:.2f}',
                    'Benchmark': '{:.2f}',
                    'Difference': '{:+.2f}' # Show sign for difference
                }).applymap(lambda x: 'color: red' if isinstance(x, (int, float)) and x > 0 else ('color: green' if isinstance(x, (int, float)) and x < 0 else None), subset=['Difference']), # Basic conditional formatting
                 use_container_width=True)
            else:
                st.warning("Could not generate comparison table.")
        else:
            st.info("Upload benchmark data to see the comparison.")


# --- Populate Tab 5: Export & Data ---
with tabs[4]:
    st.header("📤 Export & Data")
    st.markdown("_Download insights, data, or templates._")

    export_cols = st.columns(2)

    with export_cols[0]:
        st.subheader("Download Processed Data")
        if st.session_state.project_data is not None:
            st.markdown("**Full Project Data (with Predictions):**")
            df_export = st.session_state.project_data.copy()
            # Option 1: CSV
            csv_data = df_to_csv_bytes(df_export)
            st.download_button("Download Data (.csv)", csv_data, "risklens_project_data.csv", "text/csv", key="export_data_csv")
            # Option 2: Excel
            excel_data = df_to_excel_bytes({'Project Data': df_export})
            st.download_button("Download Data (.xlsx)", excel_data, "risklens_project_data.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", key="export_data_excel")
        else:
            st.info("No project data loaded.")

        # Download Clustered Data
        if st.session_state.clustering_results is not None:
            st.markdown("**Clustering Results:**")
            df_cluster_export = st.session_state.clustering_results.copy()
            csv_cluster = df_to_csv_bytes(df_cluster_export)
            st.download_button("Download Cluster Data (.csv)", csv_cluster, "risklens_cluster_data.csv", "text/csv", key="export_cluster_csv")
            excel_cluster = df_to_excel_bytes({'Clustering Results': df_cluster_export})
            st.download_button("Download Cluster Data (.xlsx)", excel_cluster, "risklens_cluster_data.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", key="export_cluster_excel")
        else:
            st.info("Run clustering first to download results.")

    with export_cols[1]:
        st.subheader("Download Model Insights")
        if st.session_state.model_trained:
            # Feature Importance
            if st.session_state.feature_importance is not None:
                st.markdown("**Feature Importance:**")
                csv_fi = df_to_csv_bytes(st.session_state.feature_importance)
                st.download_button("Download Importance (.csv)", csv_fi, "risklens_feature_importance.csv", "text/csv", key="export_fi_csv")
                excel_fi = df_to_excel_bytes({'Feature Importance': st.session_state.feature_importance})
                st.download_button("Download Importance (.xlsx)", excel_fi, "risklens_feature_importance.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", key="export_fi_excel")
            else:
                st.info("Feature importance not available.")

            # Performance Metrics (Example: from classification report)
            try:
                 y_true_export = st.session_state.y_test
                 y_pred_prob_export = st.session_state.y_pred_prob
                 y_pred_class_export = (y_pred_prob_export >= st.session_state.current_prediction_threshold).astype(int)
                 report_dict_export = classification_report(y_true_export, y_pred_class_export, output_dict=True, zero_division=0)
                 report_df_export = pd.DataFrame(report_dict_export).transpose()
                 st.markdown("**Model Performance Metrics:**")
                 csv_metrics = df_to_csv_bytes(report_df_export) # Export includes index here
                 st.download_button("Download Metrics (.csv)", csv_metrics, "risklens_model_metrics.csv", "text/csv", key="export_metrics_csv")
                 excel_metrics = df_to_excel_bytes({'Metrics': report_df_export.reset_index()}) # Reset index for excel
                 st.download_button("Download Metrics (.xlsx)", excel_metrics, "risklens_model_metrics.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", key="export_metrics_excel")
            except Exception:
                 st.info("Model metrics not available for download.")


            # Combined Insights Report (PDF)
            st.markdown("**Summary Report (PDF):**")
            current_kpis_pdf = create_kpi_summary(st.session_state.project_data, 'DerailmentRisk_Predicted_Prob', 'DerailmentRisk_Predicted')
            model_name_pdf = f"{st.session_state.selected_model}{' (AutoML)' if st.session_state.use_automl else ''}"
            pdf_bytes = export_insights_pdf(current_kpis_pdf, st.session_state.feature_importance, model_name_pdf, st.session_state.current_prediction_threshold)
            if pdf_bytes:
                 st.download_button(
                     label="Download Insights Report (.pdf)",
                     data=pdf_bytes,
                     file_name="RiskLensPro_Insights.pdf",
                     mime="application/pdf",
                     key="export_pdf_report"
                 )
            # else: Error message handled within the function if FPDF is missing or fails

            # ZIP Export (Example: Data + PDF Report)
            st.markdown("**Combined Export (ZIP):**")
            if st.session_state.project_data is not None and pdf_bytes:
                 files_to_zip = {
                     "risklens_project_data.csv": df_to_csv_bytes(st.session_state.project_data),
                     "RiskLensPro_Insights.pdf": pdf_bytes
                 }
                 # Optionally add other files like feature importance
                 if st.session_state.feature_importance is not None:
                     files_to_zip["risklens_feature_importance.csv"] = df_to_csv_bytes(st.session_state.feature_importance)

                 zip_bytes = create_download_zip(files_to_zip)
                 st.download_button(
                     label="Download Combined (.zip)",
                     data=zip_bytes,
                     file_name="RiskLensPro_Export_Package.zip",
                     mime="application/zip",
                     key="export_zip_package"
                 )
            else:
                 st.info("Requires loaded data and successful PDF report generation for ZIP export.")

        else:
            st.info("Train a model first to download model insights.")

    st.markdown("---")
    # Download Templates Section (kept from Part 2/3)
    st.subheader("Download Data Templates")
    with st.container():
        st.markdown("Download sample CSV templates for uploading data.")
        num_feat_template = st.session_state.numerical_features or SIMULATION_FEATURES_NUMERIC
        cat_feat_template = st.session_state.categorical_features or SIMULATION_FEATURES_CATEGORICAL
        required_template = list(REQUIRED_COLUMNS) + [TARGET_VARIABLE, 'ActualCost', 'ActualScheduleDays', 'CostVariancePerc', 'ScheduleVariancePerc']

        template_cols = {
            'Project Data': ['ProjectID', 'ProjectName'] + required_template + num_feat_template + cat_feat_template,
            # 'Risk Register': ['ProjectID', 'RiskID', ...], # Define if needed
            'Benchmark Data': ['MetricName', 'BenchmarkValue']
        }
        for data_type, cols in template_cols.items():
             unique_cols = pd.Index(cols).unique().tolist()
             template_df = pd.DataFrame(columns=unique_cols)
             csv_bytes = df_to_csv_bytes(template_df)
             st.download_button(
                 label=f"Download {data_type} Template (.csv)", data=csv_bytes,
                 file_name=f"{data_type.lower().replace(' ', '_')}_template.csv", mime='text/csv',
                 key=f"download_template_{data_type.lower().replace(' ', '_')}_tab_final" # Unique key
             )

# --- Final Touches & Footer ---
st.sidebar.markdown("---")
st.sidebar.info("✅ Workflow Complete")

st.markdown("---")
st.caption("RiskLens Pro - Enhanced Analytics Application")

# End of Part 4 / End of App